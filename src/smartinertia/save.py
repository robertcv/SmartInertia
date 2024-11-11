import logging
import os
import pickle
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook
from PyQt5.QtCore import QStandardPaths

from smartinertia.dialogs import RunConf

log = logging.getLogger(__name__)
DOCUMENTS_PATH = QStandardPaths.writableLocation(QStandardPaths.DocumentsLocation)
APPDATA_PATH = QStandardPaths.writableLocation(QStandardPaths.AppLocalDataLocation)
HEADER = [
    "DateTime", "Name", "Weight", "Load", "Pulley",
    "Velocity Con Max", "Velocity Ecc Max", "Velocity Con Mean", "Velocity Ecc Mean",
    "Force Con Max", "Force Ecc Max", "Force Con Mean", "Force Ecc Mean",
    "Power Con Max", "Power Ecc Max", "Power Con Mean", "Power Ecc Mean"
]


def iso2win(iso_timee: str) -> str:
    """Fix datetime in iso formated string to be used in Windows file names."""
    return iso_timee.replace(":", "-").split(".")[0]


def save_run(run_data: 'RunData', run_conf: RunConf, current_data_time: datetime):
    """Save the results of this run into a file."""
    save_dir = os.path.join(DOCUMENTS_PATH, "SmartInertia")
    save_file = os.path.join(save_dir,
                             f"measurements_{current_data_time.strftime('%Y-%m-%d')}.xlsx")
    new_row = [
        current_data_time.isoformat(), run_conf.name, run_conf.weight, run_conf.load, run_conf.pulley,
        run_data.v_con_max, run_data.v_ecc_max, run_data.v_con_mean, run_data.v_ecc_mean,
        run_data.f_con_max, run_data.f_ecc_max, run_data.f_con_mean, run_data.f_ecc_mean,
        run_data.p_con_max, run_data.p_ecc_max, run_data.p_con_mean, run_data.p_ecc_mean,
    ]

    if os.path.exists(save_file):
        try:
            wb = load_workbook(save_file)
            wb.active.append(new_row)
            wb.save(filename=save_file)
            log.info(f"Run saved to existing {save_file}.")
        except:
            log.exception("Run couldn't be saved to existing file!")
    else:
        if not os.path.exists(save_dir):
            try:
                os.mkdir(save_dir)
            except:
                log.exception("Couldn't create save directory!")

        wb = Workbook()
        sheet = wb.active
        sheet.append(HEADER)
        sheet.append(new_row)
        try:
            wb.save(filename=save_file)
            log.info(f"Run saved to new file {save_file}.")
        except:
            log.exception("Run couldn't be saved to new file!")


def save_run_more(run_datas: List['RunData'], run_conf: RunConf, current_data_time: datetime):
    """Save the more results of this run into a file."""

    save_dir = os.path.join(DOCUMENTS_PATH, "SmartInertia", f"measurements_{current_data_time.strftime('%Y-%m-%d')}")
    save_file = os.path.join(save_dir, f"{iso2win(current_data_time.isoformat())}_{run_conf.name}_{run_conf.load}.xlsx")

    if not os.path.exists(save_dir):
        try:
            os.mkdir(save_dir)
        except:
            log.exception("Couldn't create save directory!")

    wb = Workbook()
    sheet = wb.active
    sheet.append(HEADER)

    for rd in run_datas:
        new_row = [
            current_data_time.isoformat(), run_conf.name, run_conf.weight, run_conf.load, run_conf.pulley,
            rd.v_con_max, rd.v_ecc_max, rd.v_con_mean, rd.v_ecc_mean,
            rd.f_con_max, rd.f_ecc_max, rd.f_con_mean, rd.f_ecc_mean,
            rd.p_con_max, rd.p_ecc_max, rd.p_con_mean, rd.p_ecc_mean,
        ]
        sheet.append(new_row)

    try:
        wb.save(filename=save_file)
        log.info(f"Run saved to new file {save_file}.")
    except:
        log.exception("Run couldn't be saved to new file!")


def save_data(data: 'DataSet', current_data_time: datetime):
    """Save dataset to file."""
    save_data_loc = os.path.join(APPDATA_PATH, "SmartInertia", f"{iso2win(current_data_time.isoformat())}.p")
    try:
        pickle.dump({"x": data.x, "y": data.y}, open(save_data_loc, "wb"))
        log.info(f"Saved raw run to file.")
    except:
        log.exception("Cannot save run data!")
